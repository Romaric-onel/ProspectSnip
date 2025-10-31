import random
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.workbook.workbook import Workbook as Book
from requests.adapters import HTTPAdapter
from tabulate import tabulate
from urllib3.util.retry import Retry  # , RequestHistory

from utils import entry_empty, env_HEADERS

DEPARTMENT_FILE = Path("sources", "departement_fr_2025.csv")
COMMUNE_FILE = Path("sources", "commune_fr_2025.csv.csv")
ANNEE_DERNIERE = datetime.today().year - 1


HEADER_FONT = Font(
    name="Roboto", size=13, bold=True, vertAlign=None, underline="none", color="FCFFDE"
)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFC107")

ROW_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")

def rotate_header(session: requests.Session) -> requests.Session:
    global env_HEADERS
    HEADERS: dict = {"User-Agent": random.choice(env_HEADERS)}
    session.headers.update(HEADERS)
    return session


def navigate_location(loc: bool) -> int:
    if loc == True:
        filename = DEPARTMENT_FILE
        order = "REG"
        ch = "Département"
        name: list = [1, 5]
        step = 10
    else:
        filename = COMMUNE_FILE
        order = "COM"
        ch = "Commune"
        name: list = [9]
        step = 15
    data: pd.DataFrame = pd.read_csv(filename).sort_values(by=order, ascending=True)
    choice = 0

    i = 0
    print("---Bienvenue dans le menu de navigation---")
    print(
        "---Pour aller la page suivante, cliquez sur 'S' et 'P' pour la page précédente---"
    )
    while 0 <= i <= len(data):
        print(
            f"--L'objectif dans ce menu est de choisir le numéro de votre {ch} cible.\nAssurez vous qu'il est sur la page visible"
        )
        print()
        print(
            tabulate(
                data.iloc[i : i + step, name],  # type: ignore
                headers=["Num", ch],
                tablefmt="github",
                showindex=True if not loc else False,
            )
        )
        print(f"\n--- Affiché {i} à {min(i+step, len(data))} sur {len(data)} ---\n")
        cmd = input("Saisissez une commande ... ")
        if cmd.upper() == "S":
            if i == len(data) - step:
                print("Vous êtes déjà à la dernière page")
                return navigate_location(loc)
            else:
                i += step
        elif cmd.upper() == "P":
            if i == 0:
                print("Vous êtes à la première page de navigation")
                return navigate_location(loc)
            else:
                i -= step
        elif not entry_empty(cmd):
            print(
                "Vous avez entré une commande vide.\nRetour au début du menu de navigation"
            )
        else:

            try:
                choice_num = int(cmd)

            except ValueError:
                print("Commande non reconnue")
                return navigate_location(loc)
            else:
                if not i in range(i, i + step):
                    print(
                        "Vous ne pouvez entrer que les numéros de {ch} présents sur cette page.\nVous pouvez appuyer sur 'S' pour la page suivante ou 'P' pour la page précédente"
                    )
                break

    if loc:
        choice = data.loc[choice_num - 1, "DEP"]
    else:
        choice = data.loc[choice_num, "COM"]

    print(f"Excellent !!! Vous avez ajouté le {ch} {choice} à votre ciblage")
    return choice


def create_session():
    session = requests.Session()
    retries = Retry(
        total=5, backoff_factor=2, status_forcelist=(429, 500, 502, 503, 504)
    )
    session.mount("https://", HTTPAdapter(max_retries=retries))
    session.mount("http://", HTTPAdapter(max_retries=retries))
    return session
