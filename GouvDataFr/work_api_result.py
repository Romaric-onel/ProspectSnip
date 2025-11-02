from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook

from . import (
    ANNEE_DERNIERE,
    HEADER_FILL,
    HEADER_FONT,
    ID_REQUETE,
    asyncio,
    chain,
    json,
    openpyxl,
    os,
    pandas,
)
from .use_api import combine_all_result
from .utils_fr import ROW_FILL

start = 0


def resultat_to_sheet():
    global start
    raw_data = asyncio.run(combine_all_result())
    data = tuple(chain.from_iterable(item["results"] for item in raw_data))

    # création du fichier excel
    excel = openpyxl.Workbook()

    # création des feuilles sheet

    excel_presentation: Workbook = excel.active  # pyright: ignore[reportAssignmentType]
    excel_membre: Workbook = excel.create_sheet(title="Membre")
    excel_finance: Workbook = excel.create_sheet(title="Finance")

    # ....
    excel_presentation.title = (  # pyright: ignore[reportAttributeAccessIssue]
        "Présentation"
    )

    # remplir les données dans les feuilles
    for i, cible in enumerate(data):
        # données
        name = cible["nom_complet"]
        members = [member for member in cible["dirigeants"]]
        finance: dict | None = cible["finances"]

        annee = tuple(finance.keys())[0] if finance else None
        ca = finance[annee]["ca"] if finance else None
        resultat_net = finance[annee]["resultat_net"] if finance else None

        # remplissage feuille 1

        excel_presentation.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2, column=1, value=name
        )
        excel_presentation.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2, column=2, value=cible["date_creation"]
        )

        excel_presentation.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2, column=3, value="✅" if cible["date_fermeture"] else "❌"
        )
        excel_presentation.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2, column=4, value=cible["siege"]["adresse"]
        )

        # remplissage feuille 2

        if members:

            excel_membre.cell(  # pyright: ignore[reportAttributeAccessIssue]
                row=2 + start, column=1, value=name
            )
            excel_membre.merge_cells(  # pyright: ignore[reportAttributeAccessIssue]
                start_row=2 + start,
                end_row=2 + start + len(members) - 1,
                start_column=1,
                end_column=1,
            )

            num = start

            for membre in members:
                # données
                nom = membre.get("nom")
                prenoms = membre.get("prenoms")
                qualite = membre.get("qualite")
                type = membre.get("type_dirigeant")
                naissance = membre.get("annee_de_naissance")
                denomination = membre.get("denomination")
                # remplissage
                excel_membre.cell(  # pyright: ignore[reportAttributeAccessIssue]
                    row=num + 2, column=2, value=nom
                )
                excel_membre.cell(  # pyright: ignore[reportAttributeAccessIssue]
                    row=num + 2, column=3, value=prenoms
                )
                excel_membre.cell(  # pyright: ignore[reportAttributeAccessIssue]
                    row=num + 2, column=4, value=qualite
                )
                excel_membre.cell(  # pyright: ignore[reportAttributeAccessIssue]
                    row=num + 2,
                    column=5,
                    value=(
                        "✅"
                        if type == "personne physique"
                        else "❌" if type == "personne morale" else None
                    ),
                )

                excel_membre.cell(  # pyright: ignore[reportAttributeAccessIssue]
                    row=num + 2,
                    column=6,
                    value=(
                        int(naissance)
                        if naissance and naissance != "[NON-DIFFUSIBLE]"
                        else naissance
                    ),
                )
                if nom is None and prenoms is None:
                    excel_membre.cell(  # pyright: ignore[reportAttributeAccessIssue]
                        row=num + 2, column=2, value=denomination
                    )

                num += 1
        else:
            excel_membre.cell(  # pyright: ignore[reportAttributeAccessIssue]
                row=start + 2, column=2, value=None
            )
        start += len(members) + 1

        # remplissage feuille 3

        excel_finance.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2, column=1, value=name
        )
        excel_finance.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2, column=2, value=ca
        )
        excel_finance.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2, column=3, value=resultat_net
        )
        excel_finance.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2, column=4, value=int(annee) if annee else None
        )

        excel_finance.cell(  # pyright: ignore[reportAttributeAccessIssue]
            row=i + 2,
            column=5,
            value=(
                int(cible["tranche_effectif_salarie"])
                if cible["tranche_effectif_salarie"] != "NN"
                else None
            ),
        )

    # remplir les en têtes des feuilles
    header1 = (
        "Nom de la structure",
        "Date de création",
        "Est fermé ?",
        "Adresse du siège",
    )
    header2 = (
        "Nom de la structure",
        "Nom des membres / Dénomination",
        "Prénoms des membres",
        "Qualité dans la structure",
        "Personne physique ?",
        "Année de naissance",
    )
    header3 = (
        "Nom de la structure",
        "Chiffre d'affaires",
        "Résultat net",
        "Année",
        "Effectif salarié",
    )

    for sheet, head in zip(
        (excel_presentation, excel_membre, excel_finance),
        (header1, header2, header3),
        strict=True,
    ):
        for ind, cell in enumerate(sheet[1]):  # pyright: ignore[reportArgumentType]
            sheet.cell(  # pyright: ignore[reportAttributeAccessIssue]
                row=1, column=ind + 1, value=head[ind]
            )

    for worksheet in excel:
        worksheet.freeze_panes = "B2"

        for ie, row in enumerate(worksheet.rows):
            ie += 1
            if ie == 1:
                for cell in worksheet[ie]:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL

                    column_letter = get_column_letter(
                        cell.column  # pyright: ignore[reportArgumentType]
                    )
                    adjusted_width = len(str(cell.value)) + 10
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            elif not ie % 2:
                continue

            else:
                if ie == 1:
                    continue
                for cell in worksheet[ie]:
                    cell.fill = ROW_FILL

    excel.save(f"Analyse par ProspectSnip ~ {ID_REQUETE}.xlsx")
    return excel
